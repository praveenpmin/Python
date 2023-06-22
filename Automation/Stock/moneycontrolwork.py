from ast import Pass
from datetime import date
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
import datetime
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time 
#import org.openqa.selenium.JavascriptExecutor;
# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver-2") 
  
# To maximize the browser window 
browser.maximize_window()
browser.get('https://www.moneycontrol.com')
wb = openpyxl.load_workbook("example.xlsx")
sheet1 = wb.active
sheet=wb['Sheet1']
B=0
#print(sheet.max_row)
sheet.insert_cols(2)
#sheet.move_range(B)
#sheet.insert_rows(1)
wb.save("example.xlsx")
time.sleep(10)
for row in sheet['A']:
    print (row)
    print(row.value)
    time.sleep(5)
    if (row.value is not None):
        search = browser.find_element("xpath", '//*[@id="search_str"]')
        search.send_keys(row.value)
        search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
        search_Btn.click()
        #print('first')
        time.sleep(5)
        #print('sec')
        #print(row.value)
        try:
            stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
            B=B+1
            print(stk_Price)
            sheet['B'+str(B)]=stk_Price
        except NoSuchElementException:
            #if(row.value != 'date'):
                B=B+1
                sheet['B'+str(B)]='NA'
                print('NA')
                Pass
            #else:
            #    B=B+1
            #    sheet['B'+str(B)]=date
            #    print(date)
            #    Pass
    else:
        wb.save("example.xlsx")        
#print(len(sheet['B']))
#C=sheet.max_row+1
#sheet['B'+str(C)]=datetime.datetime.now()
c=B+1
sheet['B'+str(c)]=datetime.datetime.now()
wb.save("example.xlsx")
browser.quit()       