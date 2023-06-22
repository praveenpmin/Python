from ast import Pass
from datetime import date
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
import datetime
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import re
import time
from openpyxl import formatting, styles, Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
#import org.openqa.selenium.JavascriptExecutor;
# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver-2") 
  
# To maximize the browser window 
browser.maximize_window()
browser.get('https://www.moneycontrol.com')
wb = openpyxl.load_workbook("below100.xlsx")
sheet1 = wb.active
sheet=wb['Sheet1']
B=0
#print(sheet.max_row)
sheet.insert_cols(2)
#sheet.move_range(B)
#sheet.insert_rows(1)
wb.save("below100.xlsx")
time.sleep(10)
for row in sheet['A']:
    print (row)
    print(row.value)
    time.sleep(5)
    if (row.value is not None or row.value !='NA'):
        search = browser.find_element("xpath", '//*[@id="search_str"]')
        search.send_keys(row.value)
        search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
        search_Btn.click()
        #print('first')
        time.sleep(5)
        #print('sec')
        #print(row.value)
        try:
            browser.refresh()
            time.sleep(2)
            stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
            B=B+1
            print(stk_Price)
            if(row.value is not None or row.value !='NA'):
                sheet[('B'+str(B))]=re.sub('[\,]', '', stk_Price)
                #sheet['B'+str(B)]
                print(sheet['B'+str(B)])
                wb.save("below100.xlsx")
            else:
                B=B+1
                sheet['B'+str(B)]='NA'
                print('NA')
                wb.save("below100.xlsx")
                Pass
        except NoSuchElementException:
            #if(row.value != 'date'):
            B=B+1
            sheet['B'+str(B)]='NA'
            print('NA')
            wb.save("below100.xlsx")
            Pass
            #else:
            #    B=B+1
            #    sheet['B'+str(B)]=date
            #    print(date)
            #    Pass
    else:
        wb.save("below100.xlsx")        
#print(len(sheet['B']))
#C=sheet.max_row+1
#sheet['B'+str(C)]=datetime.datetime.now()
#c=B+1
#sheet['B'+str(c)]=datetime.datetime.now()
#for m in range(sheet.min_row, sheet.max_row):
    #if(int(sheet.cell(row=m,column=2).value)>int(sheet.cell(row=m,column=3).value)):
#    sec_col=sheet.cell(row=m,column=2).value
#    third_col=sheet.cell(row=m,column=3).value
 #   print(sec_col, third_col)
#    if(sec_col is not None and third_col is not None and sec_col!='NA' and third_col!='NA' and int(float(sec_col))>int(float(third_col))):
        #print(int(sheet.cell(row=m,column=2).value))
        #print(int(sheet.cell(row=m,column=3).value))
#        print ('true')
#        greenFill = PatternFill(start_color='0000FF00',
#                   end_color='0000FF00',
#                   fill_type='solid')
#        sheet.cell(row=m,column=2).fill=greenFill
#    else:
        #print(int(sheet.cell(row=m,column=2).value))
        #print(int(sheet.cell(row=m,column=3).value))
#        print('fail')
#        redFill = PatternFill(start_color='FFFF0000',
 #                  end_color='FFFF0000',
  #                 fill_type='solid')
#        sheet.cell(row=m,column=2).fill=redFill
wb.save("below100.xlsx")
browser.quit()       