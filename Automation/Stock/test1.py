from ast import Pass
#from datetime import date
from xmlrpc.client import NOT_WELLFORMED_ERROR
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from datetime import datetime
from openpyxl.styles import Font
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time
import re

def excelCreate():
    #wb = openpyxl.load_workbook("mytrade.xlsx")
    #+datetime.now().strftime("%Y%m%d")+'.xlsx'
    wb = openpyxl.Workbook() 
    ws=wb.active
    #ws1=wb['Sheet1']
    ws=wb.create_sheet('date',2)
    now = datetime.now() 
    ws.title=now.strftime('%m %d %Y, %H %M %S')
    ws['A1']='Stock Name'
    ws['B1']='Price'
    ws['C1']='Day Low'
    ws['D1']='Day High'
    ws['E1']='Year Low'
    ws['F1']='Year High'
    ws['G1']='Diff with low'
    ws['H1']='Diff with high'
    ws['I1']='Diff between low and High'
    ws['J1']='Diff between Day high and Day low'
    ws['K1']='Date'
    wb.save(filename='mytrade'+datetime.now().strftime("%Y%m%d")+'.xlsx')
    wb.remove_sheet('Sheet')
    # webdriver path set 
    #browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver") /Users/praveenbabu/Documents/python/Python/Automation
excelCreate()