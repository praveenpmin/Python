from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from selenium.webdriver.common.by import By
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time 
#import org.openqa.selenium.JavascriptExecutor;
# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver-2") 
  
# To maximize the browser window 
browser.maximize_window()
#browser.execute_script(("arguments[0].click();", loginBtn))
browser.get('https://www.moneycontrol.com')
#search = browser.find_element("xpath", '//*[@id="search_str"]')
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb.active
B=0
for row in sheet['A']:
    print (row)
    print(row.value)
    #x1 = sheet['A1']
    #print(x1.value)
#search.send_keys('HBL Power Systems')
    time.sleep(10)
    search = browser.find_element("xpath", '//*[@id="search_str"]')
    search.send_keys(row.value)
#search_Res = browser.find_element("xpath", '//*[@id="autosuggestlist"]/ul/li[1]/a/span')
#search_Res.click()
    search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
    search_Btn.click()
    print('first')
    time.sleep(30)
    print('sec')
#stk_Price = browser.find_element("xpath", '//*[@id="nsecp"]').text //*[@id="nsecp"]   data-numberanimate-value="207.20"
    stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
    #stk_Price = browser.find_element("xpath", '/html/body/header/div/div[4]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("rel")
    print('3rd') #/html/body/header/div/div[4]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2] /html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]
#stk_Price.__getattribute__("data-numberanimate-value")
#print(sheet.max_row)
#print(sheet.max_column)
    print(stk_Price)
    B=B+1
    sheet['B'+str(B)]=stk_Price
   # row in sheet['B']
    #print (row)
    #row.value=stk_Price,""
    #print (row)
    #print(row.value)
# iterate through excel and display data
#for i in range(1, sheet.max_row+1):
 #   print("\n")
   # print("Row ", i, " data :")
      
    #for j in range(1, sheet.max_column+1):
     #   cell_obj = sheet.cell(row=i, column=j)
      #  print(cell_obj.value, end=" ")
#sheet[x2].value =stk_Price
#for j in range(1, sheet.max_column+1):
 #   cell_obj = sheet.cell(row=i, column=j)
  #  #print(cell_obj.value)

#for row in sheet['A']:
 #   print(row.value)
#for row in sheet.iter_rows(1, sheet.max_row):
 #   for cell in row:
  #      print(cell.value)
wb.save('example.xlsx')

browser.quit()