from ast import Pass
from tkinter.tix import CELL
#from datetime import date
from xmlrpc.client import NOT_WELLFORMED_ERROR
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from openpyxl import formatting, styles, Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time 

wb = openpyxl.load_workbook("example1.xlsx")
sheet1 = wb.active
sheet=wb['Sheet1']
sheet=wb.active
B=0

for m in range(sheet.min_column, sheet.max_column):
    #(sheet.cell(row=m,column=2).value.isnumeric
    if(int(float(sheet.cell(row=m,column=2).value))>int(float(sheet.cell(row=m,column=3).value))):
        print(int(sheet.cell(row=m,column=2).value))
        print(int(sheet.cell(row=m,column=3).value))
        print ('true')
        greenFill = PatternFill(start_color='0000FF00',
                   end_color='0000FF00',
                   fill_type='solid')
        sheet.cell(row=m,column=2).fill=greenFill
    else:
        print(int(float(sheet.cell(row=m,column=2).value)))
        print(int(float(sheet.cell(row=m,column=3).value)))
        print('fail')
        #red_fill = PatternFill(bgColor="FFC7CE")
        #sheet.conditional_formatting.add('B1:B10', formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        #CELL.format=redFill
        #sheet.cell(row=m,column=2).fill=redFill
        #sheet.conditional_formatting.add('B1:B10', formatting.rule.CellIsRule(operator='lessThan', formula=['0'], fill=redFill))
        sheet.cell(row=m,column=2).fill=redFill

wb.save("example1.xlsx")