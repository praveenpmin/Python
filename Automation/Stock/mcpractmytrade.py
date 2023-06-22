from ast import Pass
#from datetime import date
from xmlrpc.client import NOT_WELLFORMED_ERROR
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from datetime import datetime
from openpyxl.styles import Font
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time
import re

# webdriver path set 
#browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver") 
# To maximize the browser window 
#browser.maximize_window()
#browser.get('https://www.moneycontrol.com')
wb = openpyxl.load_workbook("mytrade.xlsx")
#sheet = wb.active
ws=wb.active
A=1
G=1
B=1
C=1
D=1
E=1
F=1
H=1
I=1
J=1
K=1
#Lowandhigh=workbook['Sheet2']
#Lowandhigh=wb.active
#sheet['A10']='help'
ws1=wb['Sheet1']
ws=wb.create_sheet('date',2)
now = datetime.now() 
ws.title=now.strftime('%m %d %Y, %H %M %S')
#datetime.datetime.now
#ws.insert_cols(2)

ws['A1']='Stock Name'
ws['B1']='Price'
ws['C1']='Day Low'
ws['D1']='Day High'
ws['E1']='Year Low'
ws['F1']='Year High'
ws['G1']='Diff with low'
ws['H1']='Diff with high'
ws['I1']='Diff between low and High'
ws['J1']='Diff between Day high and Day low'
ws['K1']='Date'
wb.save("mytrade.xlsx")
# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver") 
# To maximize the browser window 
browser.maximize_window()
browser.get('https://www.moneycontrol.com')
time.sleep(2)
A=1
for row in ws1['A']:
    print (row)
    print(row.value)
    time.sleep(5)
    if (row.value is not None):
        search = browser.find_element("xpath", '//*[@id="search_str"]')
        search.send_keys(row.value)
        search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
        search_Btn.click()
        
        #print('first')
        time.sleep(5)
        #print('sec')
        #print(row.value)
        try:
            stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
            day_Low = browser.find_element("xpath", '//*[@id="sp_low"]').text
            day_High = browser.find_element("xpath", '//*[@id="sp_high"]').text
            yr_Low = browser.find_element("xpath", '//*[@id="sp_yearlylow"]').text
            yr_High = browser.find_element("xpath", '//*[@id="sp_yearlyhigh"]').text
            B=B+1
            C=C+1
            D=D+1
            E=E+1
            F=F+1
            G=G+1
            H=H+1
            I=I+1
            J=J+1
            K=K+1
            stk_Price=re.sub('[\,]', '', stk_Price)
            day_Low=re.sub('[\,]', '', day_Low)
            day_High=re.sub('[\,]', '', day_High)
            yr_Low=re.sub('[\,]', '', yr_Low)
            yr_High=re.sub('[\,]', '', yr_High)
            Diffwithlow=str(int(float(stk_Price))-int(float(day_Low)))
            Diffwithhigh=str(int(float(day_High))-int(float(stk_Price)))
            DiffbetweenlowandHigh=str(int(float(yr_High))-int(float(yr_Low)))
            DiffbetweenDayhighandyrhigh=str(int(float(day_High))-int(float(day_Low)))
            print(stk_Price,day_Low,day_Low,day_High)
            ws['A'+str(K)]=row.value
            #print(day_Low)
            ws['B'+str(B)]=stk_Price
            ws['C'+str(C)]=day_Low
            ws['D'+str(D)]=day_High
            ws['E'+str(E)]=yr_Low
            ws['F'+str(F)]=yr_High
            ws['G'+str(G)]=Diffwithlow
            ws['H'+str(H)]=Diffwithhigh
            ws['I'+str(I)]=DiffbetweenlowandHigh
            ws['J'+str(J)]=DiffbetweenDayhighandyrhigh
            ws['K'+str(K)]=datetime.now()
            wb.save("mytrade.xlsx")
        except NoSuchElementException:
            #if(row.value != 'date'):
                B=B+1
                ws['B'+str(B)]='NA'
                print('NA')
                wb.save("mytrade.xlsx")   
                Pass
            #else:
            #    B=B+1
            #    sheet['B'+str(B)]=date
            #    print(date)
            #    Pass
    else:
        wb.save("mytrade.xlsx")        
#print(len(sheet['B']))
#C=sheet.max_row+1
#sheet['B'+str(C)]=datetime.datetime.now()
#c=B+1
#ws['B'+str(c)]=datetime.now()
wb.save("mytrade.xlsx")
browser.quit()