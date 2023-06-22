from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from selenium.webdriver.common.by import By
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time 
#import org.openqa.selenium.JavascriptExecutor;
# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver-2") 
  
# To maximize the browser window 
browser.maximize_window()
#browser.execute_script(("arguments[0].click();", loginBtn))
browser.get('https://www.moneycontrol.com')
search = browser.find_element("xpath", '//*[@id="search_str"]')
wb = openpyxl.load_workbook("example.xlsx")
sheet = wb.active
x1 = sheet['A1']
print(x1.value)
#search.send_keys('HBL Power Systems')
search.send_keys(x1.value)
#search_Res = browser.find_element("xpath", '//*[@id="autosuggestlist"]/ul/li[1]/a/span')
#search_Res.click()
search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
search_Btn.click()
#stk_Price = browser.find_element("xpath", '//*[@id="nsecp"]').text //*[@id="nsecp"]   data-numberanimate-value="207.20"
stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
#stk_Price.__getattribute__("data-numberanimate-value")
#print(sheet.max_row)
#print(sheet.max_column)
print(stk_Price)
sheet['B1'].value=stk_Price
# iterate through excel and display data
for i in range(1, sheet.max_row+1):
    print("\n")
    row='A'
    p=sheet[row+i]
    print(p.value)
   # print("Row ", i, " data :")
      
    #for j in range(1, sheet.max_column+1):
     #   cell_obj = sheet.cell(row=i, column=j)
      #  print(cell_obj.value, end=" ")
#sheet[x2].value =stk_Price
for j in range(1, sheet.max_column+1):
    cell_obj = sheet.cell(row=i, column=j)
    print(cell_obj.value)
wb.save('example.xlsx')

browser.quit()