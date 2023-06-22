from ast import Pass
#from datetime import date
from xmlrpc.client import NOT_WELLFORMED_ERROR
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import time 

# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver-2") 
# To maximize the browser window 
browser.maximize_window()
#browser.get('https://www.moneycontrol.com')
wb = openpyxl.load_workbook("example.xlsx")
#sheet = wb.active
ws=wb.active
G=1
B=1
C=1
D=1
E=1
F=1
#Lowandhigh=workbook['Sheet2']
#Lowandhigh=wb.active
#sheet['A10']='help'
ws1=wb['Sheet1']
ws=wb.create_sheet('date',2)
now = datetime.now() 
ws.title=now.strftime('%m %d %Y, %H %M %S')
#datetime.datetime.now
#ws.insert_cols(2)
ws['B1']='Price'
ws['C1']='Day Low'
ws['D1']='Day High'
ws['E1']='Year Low'
ws['F1']='Year High'
wb.save("example.xlsx")
browser.get('https://www.moneycontrol.com')
time.sleep(2)
A=1
for row in ws1['A']:
    print (row)
    print(row.value)
    time.sleep(5)
    if (row.value is not None):
        search = browser.find_element("xpath", '//*[@id="search_str"]')
        search.send_keys(row.value)
        search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
        search_Btn.click()
        
        #print('first')
        time.sleep(5)
        #print('sec')
        #print(row.value)
        try:
            stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
            day_Low = browser.find_element("xpath", '//*[@id="sp_low"]').text
            day_High = browser.find_element("xpath", '//*[@id="sp_high"]').text
            yr_Low = browser.find_element("xpath", '//*[@id="sp_yearlylow"]').text
            yr_High = browser.find_element("xpath", '//*[@id="sp_yearlyhigh"]').text
            B=B+1
            C=C+1
            D=D+1
            E=E+1
            F=F+1
            G=G+1
            print(stk_Price,day_Low,day_Low,day_High)
            ws['A'+str(G)]=row.value
            #print(day_Low)
            ws['B'+str(B)]=stk_Price
            ws['C'+str(C)]=day_Low
            ws['D'+str(D)]=day_High
            ws['E'+str(E)]=yr_Low
            ws['F'+str(F)]=yr_High
        except NoSuchElementException:
            #if(row.value != 'date'):
                B=B+1
                ws['B'+str(B)]='NA'
                print('NA')
                Pass
            #else:
            #    B=B+1
            #    sheet['B'+str(B)]=date
            #    print(date)
            #    Pass
    else:
        wb.save("example.xlsx")        
#print(len(sheet['B']))
#C=sheet.max_row+1
#sheet['B'+str(C)]=datetime.datetime.now()
c=B+1
ws['B'+str(c)]=datetime.now()
wb.save("example.xlsx")
browser.quit()       


