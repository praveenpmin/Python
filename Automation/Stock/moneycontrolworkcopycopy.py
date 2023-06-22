from ast import Pass
from datetime import date
from multiprocessing.sharedctypes import Value
from numbers import Number
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
import datetime
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
#from selenium import WebElement 
# For using sleep function because selenium 
# works only when the all the elements of the 
# page is loaded. 
import re
import time
from openpyxl import formatting, styles, Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
#import org.openqa.selenium.JavascriptExecutor;
# webdriver path set 
wb = openpyxl.load_workbook("example1.xlsx")
sheet1 = wb.active
sheet=wb['Sheet1']
sheet = wb.active
for i in range(11, sheet.max_row+1):
    #print("\n")
    #print("Row ", i, " data :")

    for j in range(2, sheet.max_column+1):
        cell_obj = sheet.cell(row=i, column=j)
        mylist=[]
        #print(i,j)
        #list=cell_obj.value
       # print(mylist)
        mylist=cell_obj.value
        print(mylist)
    #print(max(mylist))
    largest_number = mylist[0]
    for number in mylist:
        if number > largest_number:
            largest_number = number
    print(largest_number)

        #print(cell_obj.value, end=" ")
        #print("The max value:", max((cell_obj.value)))
        #print("\n")

wb.save("example1.xlsx")
    