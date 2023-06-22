from ast import Pass
#from datetime import date
from xmlrpc.client import NOT_WELLFORMED_ERROR
from selenium import webdriver 
#from selenium.webdriver.common.keys import Keys
import openpyxl
from openpyxl import workbook
from datetime import datetime
from openpyxl.styles import Font
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time
import re
import calendar
import datetime
from datetime import datetime
import os.path
import mon

G=1
B=1
C=1
D=1
E=1
F=1
H=1
I=1
J=1
K=1
L=1
today = datetime.now()
    
#print("Current Date and Time :", today)
#print("Current Month :", today.month)
datetime_object = datetime.strptime(str(today.month), "%m")
month_name = datetime_object.strftime("%b")
#print("Short name: ",month_name)

file_Name='mytrade'+str(datetime.now())+'.xlsx'
#print(file_Name)
now = datetime.now() 
sheet_Name=now.strftime('%m %d %Y, %H %M %S')
#print(sheet_Name)

def folder_Exist():  
   
    file_Path='/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name
    isdir = os.path.isdir(file_Path)
    #print(isdir)  

    try:
        os.makedirs(f'/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name)
    except FileExistsError:
        # directory already exists
        pass
folder_Exist()

#file_Name='mytrade'+datetime.now().strftime("%Y%m%d")+'.xlsx'

def excelCreate():
    wb = openpyxl.Workbook() 
    ws=wb.active
    #ws1=wb['Sheet1']
    ws=wb.create_sheet('date',2)
    now = datetime.now() 
    #ws.title=now.strftime('%m %d %Y, %H %M %S')
    ws.title=sheet_Name
    ws['A1']='Stock Id'
    ws['B1']='Stock Name'
    ws['C1']='Price'
    ws['D1']='Day Low'
    ws['E1']='Day High'
    ws['F1']='Year Low'
    ws['G1']='Year High'
    ws['H1']='Diff with low'
    ws['I1']='Diff with high'
    ws['J1']='Diff between low and High'
    ws['K1']='Diff between Day high and Day low'
    ws['L1']='Date'
    #file_Name='mytrade'+datetime.now().strftime("%Y%m%d")+'.xlsx'
    wb.save(filename='/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)
excelCreate()

wb = openpyxl.load_workbook('/Users/praveenbabu/Documents/python/Python/Automation/Stock/mytrade.xlsx')
#print(wb.sheetnames)
wb1 = openpyxl.load_workbook('/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)
#sheet = wb.active
ws1=wb.active
ws1=wb['Sheet1']
ws2=wb1.active
ws2=wb1[sheet_Name]
# webdriver path set 
browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver") 
# To maximize the browser window 
browser.maximize_window()
browser.get('https://www.moneycontrol.com')
time.sleep(2)
A=1
for row in ws1['A']:
    print (row.value)
    #print(row)
    time.sleep(5)
    if (row.value is not None):
        search = browser.find_element("xpath", '//*[@id="search_str"]')
        search.send_keys(row.value)
        search_Btn = browser.find_element("xpath", '//*[@id="Capa_1"]')
        search_Btn.click()
        
        #print('first')
        time.sleep(5)
        #print('sec')
        #print(row.value)
        try:
            stk_Name = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[2]/div[2]/div/h1').text
            stk_Price = browser.find_element("xpath", '/html/body/div[10]/div[1]/div[4]/div[1]/div/div[1]/div/div[1]/div[2]').get_attribute("data-numberanimate-value")
            day_Low = browser.find_element("xpath", '//*[@id="sp_low"]').text
            day_High = browser.find_element("xpath", '//*[@id="sp_high"]').text
            yr_Low = browser.find_element("xpath", '//*[@id="sp_yearlylow"]').text
            yr_High = browser.find_element("xpath", '//*[@id="sp_yearlyhigh"]').text
            B=B+1
            C=C+1
            D=D+1
            E=E+1
            F=F+1
            G=G+1
            H=H+1
            I=I+1
            J=J+1
            K=K+1
            L=L+1
            #M=M+1
            stk_Price=re.sub('[\,]', '', stk_Price)
            day_Low=re.sub('[\,]', '', day_Low)
            day_High=re.sub('[\,]', '', day_High)
            yr_Low=re.sub('[\,]', '', yr_Low)
            yr_High=re.sub('[\,]', '', yr_High)
            Diffwithlow=str(int(float(stk_Price))-int(float(day_Low)))
            Diffwithhigh=str(int(float(day_High))-int(float(stk_Price)))
            DiffbetweenlowandHigh=str(int(float(yr_High))-int(float(yr_Low)))
            DiffbetweenDayhighandyrhigh=str(int(float(day_High))-int(float(day_Low)))
            print(stk_Name,stk_Price,day_Low,day_Low,day_High)
            ws2['A'+str(K)]=row.value
            #ws2['B'+str[M]]=row.value
            #print(day_Low)
            ws2['B'+str(B)]=stk_Name
            ws2['C'+str(C)]=stk_Price
            ws2['D'+str(D)]=day_Low
            ws2['E'+str(E)]=day_High
            ws2['F'+str(F)]=yr_Low
            ws2['G'+str(G)]=yr_High
            ws2['H'+str(H)]=Diffwithlow
            ws2['I'+str(I)]=Diffwithhigh
            ws2['J'+str(J)]=DiffbetweenlowandHigh
            ws2['K'+str(K)]=DiffbetweenDayhighandyrhigh
            ws2['L'+str(L)]=datetime.now()
            wb1.save('/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)   
        except NoSuchElementException:
            #if(row.value != 'date'):
                B=B+1
                ws2['B'+str(B)]='NA'
                print('NA')
                wb1.save('/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)     
                Pass
    else:
        wb1.save('/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)        
wb1.remove(wb1['Sheet'])
wb1.save('/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)
browser.quit()