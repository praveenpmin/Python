import time
#from tkinter import BROWSE
from selenium import webdriver 
import openpyxl
def fileRead():
  
    # load excel with its path
    wrkbk = openpyxl.load_workbook('/Users/praveenbabu/Documents/python/Python/Automation/Stock/mytrade.xlsx')
    
    sh = wrkbk.active
    
    # iterate through excel and display data
    for row in sh.iter_rows(min_row=1, min_col=1, max_row=12, max_col=3):
        for cell in row:
            print(cell.value, end=" ")
        print()

fileRead()
'''wb = openpyxl.load_workbook('/Users/praveenbabu/Documents/python/Python/Automation/Stock/mytrade.xlsx')
    #print(wb.sheetnames)
    wb1 = openpyxl.load_workbook('/Users/praveenbabu/Documents/python/Python/Automation/Practice/'+month_name+'/'+file_Name)
    #sheet = wb.active
    ws1=wb.active
    ws1=wb['Sheet1']
    ws2=wb1.active
    ws2=wb1[sheet_Name]
    # webdriver path set 
    #browser = webdriver.Chrome("/Users/praveenbabu/Documents/python/Python/Automation/chromedriver") 
    # To maximize the browser window 
    #browser.maximize_window()
    #browser.get('https://www.moneycontrol.com')
    #time.sleep(2)
    A=1
    for row in ws1['A']:
        print (row.value)
        #print(row)
        time.sleep(5)
        if (row.value is not None):
            search = browser.find_element("xpath", '//*[@id="search_str"]')
            search.send_keys(row.value)
            search_Btn = BROWSE.find_element("xpath", '//*[@id="Capa_1"]')
            search_Btn.click()
            
            #print('first')
            time.sleep(5)
            #print('sec')'''