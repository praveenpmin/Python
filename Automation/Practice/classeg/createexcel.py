import datetime
import openpyxl
import time
from datetime import datetime

def excelCreate():
    today = datetime.now()
    now = datetime.now()
    datetime_object = datetime.strptime(str(today.month), "%m")
    month_name = datetime_object.strftime("%b")
    #print("Short name: ",month_name)
    file_Name='mytrade'+str(datetime.now())+'.xlsx'
    sheet_Name=now.strftime('%m %d %Y, %H %M %S')
    wb = openpyxl.Workbook() 
    ws=wb.active
    #ws1=wb['Sheet1']
    ws=wb.create_sheet('date',2)
    now = datetime.now() 
    #ws.title=now.strftime('%m %d %Y, %H %M %S')
    ws.title=sheet_Name
    ws['A1']='Stock Id'
    #ws['B1']='Stock Name'
    ws['B1']='Price'
    ws['C1']='Day Low'
    ws['D1']='Day High'
    ws['E1']='Year Low'
    ws['F1']='Year High'
    ws['G1']='Diff with low'
    ws['H1']='Diff with high'
    ws['I1']='Diff between low and High'
    ws['J1']='Diff between Day high and Day low'
    ws['K1']='Date'
    #file_Name='mytrade'+datetime.now().strftime("%Y%m%d")+'.xlsx'
    wb.save(filename='/Users/praveenbabu/Documents/python/Python/Automation/Practice/classeg/'+file_Name)
excelCreate()