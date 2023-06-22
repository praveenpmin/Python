# importing openpyxl module
import openpyxl
 
# Give the location of the file
path = '/Users/praveenbabu/Documents/python/Python/Automation/Practice/mytrade.xlsx'
 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet1 = wb_obj["Input"]
#sheet2 = wb_obj["Sheet2"]
maxr = sheet1.max_row
maxc = sheet1.max_column
for r in range (1, maxr + 1):
    for c in range (1, maxc + 1):
        #sheet2.cell(row=r,column=c).value = sheet1.cell(row=r,column=c).value
        print(sheet1.cell(row=r,column=c).value)